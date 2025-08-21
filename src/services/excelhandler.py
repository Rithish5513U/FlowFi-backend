from openpyxl import load_workbook
from extensions import db
from flask_jwt_extended import get_jwt_identity

transactions = db['transactions']

class ExcelHandler:
    def __init__(self):
        self.required_columns = ['date', 'description', 'withdrawals', 'deposits', 'balance']

    def process_uploaded_file(self, file_stream):
        """
        Function to process a single uploaded Excel file (in-memory) and store unique records to DB
        Inputs:
            file_stream -> BytesIO : Uploaded file stream from client
        Outputs:
            Tuple : (List of user's transactions in dict format, status code)
        """
        try:
            # Read Excel file
            wb = load_workbook(file_stream, data_only=True)
            sheet = wb.active

            # Extract headers
            headers = [str(cell.value).lower() if cell.value else "" for cell in sheet[1]]

            if not all(col in headers for col in self.required_columns):
                return {"error": "Invalid Excel format"}, 400

            # Map header name -> column index
            col_map = {col: headers.index(col) for col in self.required_columns}

            # Build "df-like" list of dicts
            records = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                record = {col: (row[col_map[col]] if row[col_map[col]] is not None else 0)
                          for col in self.required_columns}
                records.append(record)

            # Drop duplicates (like df.drop_duplicates)
            records = [dict(t) for t in {tuple(d.items()) for d in records}]

            user_email = get_jwt_identity()
            existing_entry = transactions.find_one({"email": user_email})

            if existing_entry:
                # mimic existing_df + concat + drop_duplicates
                existing_data = existing_entry.get("data", [])
                combined_data = existing_data + records
                combined_data = [dict(t) for t in {tuple(d.items()) for d in combined_data}]

                transactions.update_one(
                    {"email": user_email},
                    {"$set": {"data": combined_data}}
                )
            else:
                combined_data = records
                transactions.insert_one({
                    "email": user_email,
                    "data": combined_data
                })

            return combined_data, 200

        except Exception as e:
            return {"error": str(e)}, 500
